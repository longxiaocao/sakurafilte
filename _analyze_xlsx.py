"""
深度分析 新思路.xlsx 的真实结构
"""
import pandas as pd
import openpyxl
import re
from collections import Counter

FILE = r"d:\projects\sakurafilter\新思路.xlsx"

# 1) 加载全部 sheet
wb = openpyxl.load_workbook(FILE, data_only=True)
print("=" * 80)
print(f"[1] Sheet 总览: {wb.sheetnames}")
print(f"    共 {len(wb.sheetnames)} 个 sheet")
print("=" * 80)

# 2) 对每个 sheet 做体检
for sn in wb.sheetnames:
    ws = wb[sn]
    df = pd.read_excel(FILE, sheet_name=sn, header=None)
    print(f"\n--- Sheet: {sn} ---")
    print(f"    尺寸: {df.shape[0]} 行 x {df.shape[1]} 列")
    print(f"    前 3 行预览:")
    for i in range(min(3, len(df))):
        row = df.iloc[i].tolist()
        # 截断长字符串
        row = [str(x)[:30] + "..." if isinstance(x, str) and len(str(x)) > 30 else x for x in row]
        print(f"      R{i+1}: {row}")
